#!/usr/bin/env python3

import sys
import argparse
import json
import pandas as pd
import smtplib
import ssl
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
# from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from pathlib import Path
from sqlalchemy import create_engine
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

version = "0.1"


def get_frame(**node):
    dbConnection = None
    datas = None

    try:
        sqlEngine = create_engine('mysql+pymysql://' + node['username'] + ':' + node['password'] +
                                  '@' + node['hostname'] + ':' + str(node['port']) + '/' + node['database'],
                                  pool_recycle=3600)
        dbConnection = sqlEngine.connect()
        datas = pd.read_sql(node['query'], dbConnection)
        dbConnection.close()

    except Exception as err:
        dbConnection.close()
        print(str(err))

    return datas


parser = argparse.ArgumentParser(description='Sample Reporting tool')
parser.add_argument('--conf', type=str, help='configuration file to use (Mandatory)', dest="conf")
parser.add_argument('--dryrun', help='do not create snow ticket just add a comment', dest="dryrun", action="store_true")
parser.add_argument('--debug', help='display more info', dest="debug", action="store_true")
parser.add_argument('--version', help='display software info', dest="version", action="store_true")
parser.add_argument('--query', help='run a predefined query (Mandatory)', dest="query", type=str)
parser.add_argument('--filter', help='add an extra sql filter(Optional)', dest="filter", type=str)
parser.add_argument('--limit_rows', help='limit the results returned (Optional)', dest="limit_rows", type=str)
parser.add_argument('--limit_start', help='return the results starting from(Optional but require --limit_rows)',
                    dest="limit_start", type=str)

args, unknown = parser.parse_known_args()
fill = PatternFill("solid", fgColor="DDDDDD")

if len(sys.argv) == 1 or not args.conf:
    parser.print_help()
    exit()

if args.version:
    print(f"Version: {version}")
    exit()

if not os.path.exists('../conf/' + args.conf):
    print(f"Sorry the configuration file specified does not exist")
    exit()

if not args.query:
    print("Sorry a query is required")
    parser.print_help()
    exit()

with open('../conf/' + args.conf) as fconfig:
    config = json.load(fconfig)

if args.query not in config['queries'].keys():
    print(f"Sorry the query {args.query} does not exist in the configuration file")
    exit()

#
# Define input / output files, directories for templates and data results
#
outputDir = Path(config['output_directory'] + "/")
outputFile = outputDir / (config['queries'][args.query]['output_filename'] + "." +
                          config['queries'][args.query]['output_type'])
sqlFolder = Path(config['sql_directory'] + "/")
sqlFile = open(sqlFolder / config['queries'][args.query]['query_file'], "r")
sqlQuery = sqlFile.read()

if args.debug:
    print(f"Run specific query {args.query}")
    print(f"File to run : {sqlFile}")
    print(f"output file: {outputFile}")
    print(f"File to save results: {outputFile}")
    print(f"Configuration file: {args.conf}")

df_all_rows = None

if args.filter:
    if sqlQuery.find('WHERE ') != -1:
        if args.debug:
            print("Found a WHERE condition in Query")
        sqlQuery += f"\n  AND {args.filter}"
    else:
        if args.debug:
            print("Not found a WHERE condition in Query")
        sqlQuery += f"WHERE {args.filter}"

if args.limit_rows:
    if args.limit_start:
        sqlQuery += f"\nLIMIT {args.limit_rows}, {args.limit_start}"
    else:
        sqlQuery += f"\nLIMIT {args.limit_rows}"

if args.debug:
    print(f"SQL : {sqlQuery}")

for source in config['queries'][args.query]['sources']:
    if args.debug:
        print("source: %s" % source)

    if source not in config['datasources'].keys():
        print("Sorry the query %s refer to datasource %s that does not exist" % (args.query, source))
        exit()

    frame = get_frame(
        username=config['datasources'][source]['username'],
        password=config['datasources'][source]['password'],
        port=config['datasources'][source]['port'],
        database=config['datasources'][source]['database'],
        hostname=config['datasources'][source]['hostname'],
        query=sqlQuery
    )

    if 'df_all_rows' not in locals():
        df_all_rows = pd.concat([frame])
    else:
        df_all_rows = pd.concat([df_all_rows, frame])

for field in config['queries'][args.query]['fixfields']:
    if args.debug:
        print(f"field to cleanup from html stuff and fill undef value to blank: {field}")

    df_all_rows[field] = df_all_rows[field].fillna('')
    df_all_rows[field] = df_all_rows[field].apply(lambda x: BeautifulSoup(x, "lxml").get_text())

if 'sort_field' in config['queries'][args.query].keys():
    if args.debug:
        print(f"Will sort result with field : {config['queries'][args.query]['sort_field']}")
    df_all_rows.sort_values(config['queries'][args.query]['sort_field'], inplace=True)

if config['queries'][args.query]['output_type'] == "xlsx":
    if 'excel_options' in config['queries'][args.query].keys():
        if config['queries'][args.query]['excel_options']['template_file']:

            template = Path(config['template_directory'] + "/") / \
                       config['queries'][args.query]['excel_options']['template_file']

            if os.path.exists(template):
                wb = load_workbook(template)
                wb.save(outputFile)
            else:
                print(f"Sorry the excel template {template} described in {args.query} does not exist")
                exit()

    with pd.ExcelWriter(outputFile, mode='a', if_sheet_exists='overlay') as writer:
        df_all_rows.to_excel(
            writer, sheet_name=config['queries'][args.query]['excel_options']['sheet_name'],
            header=False, index=False,
            startrow=config['queries'][args.query]['excel_options']['startrow']
        )

    wb = load_workbook(outputFile)
    ws = wb[config['queries'][args.query]['excel_options']['sheet_name']]

    #
    # this section allow to put color on rows every 2 rows.
    # It takes a lot of time to do ~ 1m for 35 columon on 8.2K rows

    # for x in range(config['queries'][args.query]['excel_options']['startrow'], ws.max_row):
    #         if x % 2 != 0:
    #             for col in range(1, ws.max_column+1):
    #                 c = ws.cell(row=x, column=col)
    #                 c.fill = fill
    # wb.save(output_file)

elif config['queries'][args.query]['output_type'] == "csv":
    df_all_rows.to_csv(outputFile, index=False)

elif config['queries'][args.query]['output_type'] == "html":
    df_all_rows.to_html(outputFile)

elif config['queries'][args.query]['output_type'] == "json":
    df_all_rows.to_json(outputFile, orient='records')

else:
    print(f"sorry output_type {config['queries'][args.query]['output_type']} in related to query {args.query}",
          " is unknown")
    exit()

if "email" in config['queries'][args.query].keys():

    server = None

    if args.debug:
        print(f"Email will be sent to {config['queries'][args.query]['email']['to']}")

    #
    # Prepare email
    #
    message = MIMEMultipart()
    message["From"] = config['queries'][args.query]['email']['from']
    message["To"] = config['queries'][args.query]['email']['to']
    message["Subject"] = config['queries'][args.query]['email']['subject']

    # Not yet used
    # message["Bcc"] = receiver_email

    # Add body to email
    message.attach(MIMEText(config['queries'][args.query]['email']['body'], "plain"))

    filename = outputFile  # In same directory as script

    with open(filename, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    encoders.encode_base64(part)

    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    message.attach(part)
    text = message.as_string()

    context = ssl.create_default_context()
    try:
        server = smtplib.SMTP(config['smtp_relay']['host'], config['smtp_relay']['port'])
        server.starttls(context=context)  # Secure the connection
        server.login(config['smtp_relay']['username'], config['smtp_relay']['password'])
        server.sendmail(
            config['queries'][args.query]['email']['from'],
            config['queries'][args.query]['email']['to'],
            text
        )
    except Exception as e:
        # Print any error messages to stdout
        print(f"Error :{e}")
    finally:
        server.quit()
